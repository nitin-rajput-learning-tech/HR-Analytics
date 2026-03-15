from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from hr_analytics.exports import dataframe_to_csv_bytes, dataframe_to_excel_bytes


def test_export_payloads_include_metadata():
    df = pd.DataFrame([{"employee_number": "001", "full_name": "Example Person"}])

    csv_payload = dataframe_to_csv_bytes(
        df,
        snapshot_date="2026-03-06",
        source_file="sample.xlsx",
        note="Snapshot export",
    ).decode("utf-8")
    xlsx_payload = dataframe_to_excel_bytes(
        df,
        snapshot_date="2026-03-06",
        source_file="sample.xlsx",
        note="Snapshot export",
    )

    assert "export_snapshot_date" in csv_payload
    assert "sample.xlsx" in csv_payload
    assert len(xlsx_payload) > 0


def test_excel_export_supports_extra_sheets():
    df = pd.DataFrame([{"employee_number": "001", "full_name": "Example Person"}])
    actions_df = pd.DataFrame([{"severity": "High", "category": "Forecast"}])
    hotspots_df = pd.DataFrame([{"department": "Ops", "risk_rate": 0.25}])

    payload = dataframe_to_excel_bytes(
        df,
        snapshot_date="2026-03-06",
        source_file="sample.xlsx",
        note="Predictive export",
        summary_name="PredictiveForecast",
        extra_sheets={
            "PrioritizedActions": actions_df,
            "DepartmentHotspots": hotspots_df,
        },
    )

    workbook = load_workbook(BytesIO(payload))

    assert "PredictiveForecast" in workbook.sheetnames
    assert "Metadata" in workbook.sheetnames
    assert "PrioritizedActions" in workbook.sheetnames
    assert "DepartmentHotspots" in workbook.sheetnames
