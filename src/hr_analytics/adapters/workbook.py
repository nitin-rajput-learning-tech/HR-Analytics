from __future__ import annotations

import hashlib
import os
import json
import re
from datetime import date, datetime
from getpass import getuser
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .base import DataSourceAdapter
from ..constants import (
    COMPATIBLE_REQUIRED_FIELDS,
    EMPLOYEE_COLUMNS,
    FULL_COMPATIBILITY_FIELDS,
    MANUAL_UPLOAD_ALLOWED_LEVELS,
    OPTIONAL_COMPATIBILITY_FIELDS,
    PARTIAL_MIN_FIELDS,
    resource_path,
)
from ..models import IngestionResult, SnapshotCandidate, ValidationIssue
from ..normalization import normalize_blankable_text, normalize_display_value
from ..repository import Repository
from ..workspace import AppWorkspace

EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
MONTH_PATTERN = re.compile(
    r"(?P<day>\d{1,2})(?:st|nd|rd|th)?[\s\-]*(?:day)?[\s\-]*(?:of)?[\s\-]*(?P<month>jan|january|feb|february|mar|march|apr|april|may|jun|june|jul|july|aug|august|sep|sept|september|oct|october|nov|november|dec|december)[\s'\-]*(?P<year>\d{2,4})?",
    re.IGNORECASE,
)
MONTH_NAME_TO_NUMBER = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


class WorkbookUploadAdapter(DataSourceAdapter):
    def __init__(self, workspace: AppWorkspace, repository: Repository) -> None:
        self.workspace = workspace
        self.repository = repository
        self.column_mappings = json.loads(
            resource_path("config", "column_mappings.json").read_text(encoding="utf-8")
        )

    def ingest_snapshot(self, source: Path, *, force_reimport: bool = False) -> IngestionResult:
        raw_bytes = source.read_bytes()
        candidate = self._build_candidate(
            source_name=source.name,
            raw_bytes=raw_bytes,
            source_path=source,
            uploaded_by=resolve_uploaded_by(source_path=source),
            previous_date=None,
            override_as_of_date=None,
        )
        return self._persist_candidate(candidate, force_reimport=force_reimport)

    def review_uploaded_file(
        self,
        *,
        file_name: str,
        raw_bytes: bytes,
        override_as_of_date: date | None = None,
    ) -> SnapshotCandidate:
        return self._build_candidate(
            source_name=file_name,
            raw_bytes=raw_bytes,
            source_path=None,
            uploaded_by=resolve_uploaded_by(source_path=None),
            previous_date=None,
            override_as_of_date=override_as_of_date,
        )

    def ingest_uploaded_file(
        self,
        *,
        file_name: str,
        raw_bytes: bytes,
        override_as_of_date: date | None = None,
        force_reimport: bool = False,
    ) -> IngestionResult:
        candidate = self._build_candidate(
            source_name=file_name,
            raw_bytes=raw_bytes,
            source_path=None,
            uploaded_by=resolve_uploaded_by(source_path=None),
            previous_date=None,
            override_as_of_date=override_as_of_date,
        )
        return self._persist_candidate(candidate, force_reimport=force_reimport)

    def bootstrap_from_fixtures(self, *, force_reimport: bool = False) -> list[IngestionResult]:
        fixture_dir = self.workspace.fixture_dir()
        if not fixture_dir:
            return []
        results: list[IngestionResult] = []
        previous_date: date | None = None
        for source in self.list_bootstrap_sources():
            candidate = self._build_candidate(
                source_name=source.name,
                raw_bytes=source.read_bytes(),
                source_path=source,
                uploaded_by="fixture_bootstrap",
                previous_date=previous_date,
                override_as_of_date=None,
            )
            if candidate.as_of_date:
                previous_date = candidate.as_of_date
            results.append(self._persist_candidate(candidate, force_reimport=force_reimport))
        return results

    def list_bootstrap_sources(self) -> list[Path]:
        fixture_dir = self.workspace.fixture_dir()
        if not fixture_dir:
            return []
        return sorted(fixture_dir.glob("*.xlsx"), key=_bootstrap_sort_key)

    def list_snapshots(self) -> pd.DataFrame:
        return self.repository.list_snapshots()

    def refresh_metrics(self) -> None:
        self.repository.refresh_materializations()

    def get_validation_report(self, snapshot_id: str | None = None) -> pd.DataFrame:
        return self.repository.get_validation_report(snapshot_id)

    def _build_candidate(
        self,
        *,
        source_name: str,
        raw_bytes: bytes,
        source_path: Path | None,
        uploaded_by: str,
        previous_date: date | None,
        override_as_of_date: date | None,
    ) -> SnapshotCandidate:
        workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
        detected_sheet, header_row, mapped_headers = self._detect_employee_sheet(workbook)
        parsed_date, confidence, parse_note = parse_as_of_date(
            source_name, previous_date=previous_date
        )
        if override_as_of_date:
            parsed_date = override_as_of_date
            confidence = 1.0
            parse_note = "As-of date confirmed by user override."

        snapshot_id = hashlib.sha1(raw_bytes).hexdigest()[:16]

        if not detected_sheet or not header_row:
            return SnapshotCandidate(
                source_name=source_name,
                source_path=source_path,
                raw_bytes=raw_bytes,
                uploaded_by=uploaded_by,
                as_of_date=parsed_date,
                parse_confidence=confidence,
                parse_note=parse_note,
                detected_sheet=None,
                header_row=None,
                available_columns=[],
                missing_columns=sorted(EMPLOYEE_COLUMNS),
                compatibility_level="rejected",
                row_count=0,
                status="rejected",
                notes=["Could not detect a usable employee sheet."],
                validation_issues=[
                    ValidationIssue(
                        severity="error",
                        issue_type="sheet_detection",
                        field_name="sheet",
                        issue_count=1,
                        message="No sheet contained enough mappable employee columns.",
                    )
                ],
                snapshot_id=snapshot_id,
            )

        rows, available_columns, invalid_dates = self._extract_rows(
            workbook=workbook,
            sheet_name=detected_sheet,
            header_row=header_row,
            mapped_headers=mapped_headers,
        )
        dataframe = pd.DataFrame(rows, columns=EMPLOYEE_COLUMNS)
        compatibility_level = determine_compatibility(available_columns)
        status = "imported" if compatibility_level != "rejected" and parsed_date else "rejected"

        notes = [parse_note]
        if compatibility_level == "partial":
            notes.append("Imported with partial source coverage; some dashboard modules will show warnings.")
        if compatibility_level == "compatible_with_warnings":
            notes.append("Imported with optional field gaps; metrics remain available.")
        if not parsed_date:
            notes.append("Could not infer snapshot date from filename. Confirm the as-of date before importing.")
        missing_columns = sorted(set(EMPLOYEE_COLUMNS) - set(available_columns))
        validation_issues = self._build_validation_issues(
            dataframe=dataframe,
            invalid_dates=invalid_dates,
            available_columns=available_columns,
        )

        return SnapshotCandidate(
            source_name=source_name,
            source_path=source_path,
            raw_bytes=raw_bytes,
            uploaded_by=uploaded_by,
            as_of_date=parsed_date,
            parse_confidence=confidence,
            parse_note=parse_note,
            detected_sheet=detected_sheet,
            header_row=header_row,
            available_columns=sorted(available_columns),
            missing_columns=missing_columns,
            compatibility_level=compatibility_level,
            row_count=len(dataframe),
            status=status,
            notes=notes,
            dataframe=dataframe,
            validation_issues=validation_issues,
            snapshot_id=snapshot_id,
        )

    def _persist_candidate(
        self, candidate: SnapshotCandidate, *, force_reimport: bool = False
    ) -> IngestionResult:
        candidate.raw_file_path = self._copy_raw_file(candidate)
        snapshot_id, status = self.repository.save_snapshot(candidate, force_reimport=force_reimport)
        message = build_status_message(status, candidate)
        return IngestionResult(
            snapshot_id=snapshot_id,
            status=status,
            message=message,
            compatibility_level=candidate.compatibility_level,
            row_count=candidate.row_count,
            as_of_date=candidate.as_of_date,
            raw_file_path=candidate.raw_file_path,
            uploaded_by=candidate.uploaded_by,
            validation_issues=candidate.validation_issues,
        )

    def _copy_raw_file(self, candidate: SnapshotCandidate) -> Path:
        safe_name = sanitize_filename(candidate.source_name)
        date_part = candidate.as_of_date.isoformat() if candidate.as_of_date else "undated"
        raw_path = self.workspace.raw_uploads_dir / f"{date_part}_{candidate.snapshot_id}_{safe_name}"
        if not raw_path.exists() and candidate.raw_bytes is not None:
            raw_path.write_bytes(candidate.raw_bytes)
        return raw_path

    def _detect_employee_sheet(
        self, workbook: Any
    ) -> tuple[str | None, int | None, dict[str, str | None]]:
        best_score: tuple[int, int, int] = (-1, -1, -1)
        best_sheet: str | None = None
        best_header_row: int | None = None
        best_mapped_headers: dict[str, str | None] = {}

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            upper_bound = min(10, worksheet.max_row)
            for row_index in range(1, upper_bound + 1):
                header_values = _safe_row_values(worksheet, row_index)
                if header_values is None:
                    continue
                mapped_headers = {}
                unique_fields: set[str] = set()
                for raw_header in header_values:
                    if raw_header is None:
                        continue
                    raw_text = str(raw_header).strip()
                    mapped = self.column_mappings.get(raw_text)
                    mapped_headers[raw_text] = mapped
                    if mapped in EMPLOYEE_COLUMNS or mapped == "fallback_email":
                        unique_fields.add("work_email" if mapped == "fallback_email" else mapped)
                score = (
                    len(unique_fields),
                    1 if "employee report" in sheet_name.lower() else 0,
                    int(worksheet.max_row or 0),
                )
                if score > best_score:
                    best_score = score
                    best_sheet = sheet_name
                    best_header_row = row_index
                    best_mapped_headers = mapped_headers

        if best_score[0] < 4:
            return None, None, {}
        return best_sheet, best_header_row, best_mapped_headers

    def _extract_rows(
        self,
        *,
        workbook: Any,
        sheet_name: str,
        header_row: int,
        mapped_headers: dict[str, str | None],
    ) -> tuple[list[dict[str, Any]], set[str], dict[str, int]]:
        worksheet = workbook[sheet_name]
        header_values = _safe_row_values(worksheet, header_row)
        if header_values is None:
            return [], set(), {"date_joined": 0, "last_working_day": 0, "exit_requested_on": 0}
        headers = [str(value).strip() if value is not None else "" for value in header_values]
        available_columns = {
            mapped if mapped != "fallback_email" else "work_email"
            for mapped in mapped_headers.values()
            if mapped in EMPLOYEE_COLUMNS or mapped == "fallback_email"
        }
        rows: list[dict[str, Any]] = []
        invalid_dates = {"date_joined": 0, "last_working_day": 0, "exit_requested_on": 0}

        for values in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            canonical_row = {field: None for field in EMPLOYEE_COLUMNS}
            row_has_data = False
            for header, value in zip(headers, values):
                if not header:
                    continue
                mapped = self.column_mappings.get(header)
                if mapped == "fallback_email":
                    normalized = normalize_value("work_email", value)
                    if normalized and not canonical_row["work_email"]:
                        canonical_row["work_email"] = normalized
                        row_has_data = True
                    continue
                if mapped not in EMPLOYEE_COLUMNS:
                    continue
                normalized = normalize_value(mapped, value)
                if mapped in {"date_joined", "last_working_day", "exit_requested_on"} and value not in (None, "", " ") and normalized is None:
                    invalid_dates[mapped] += 1
                canonical_row[mapped] = normalized
                if normalized not in (None, ""):
                    row_has_data = True

            if not row_has_data:
                continue
            if not canonical_row["employee_number"] and not canonical_row["full_name"]:
                continue
            rows.append(canonical_row)

        return rows, available_columns, invalid_dates

    def _build_validation_issues(
        self,
        *,
        dataframe: pd.DataFrame,
        invalid_dates: dict[str, int],
        available_columns: set[str],
    ) -> list[ValidationIssue]:
        issues: list[ValidationIssue] = []
        if dataframe.empty:
            issues.append(
                ValidationIssue(
                    severity="error",
                    issue_type="empty_sheet",
                    field_name="rows",
                    issue_count=1,
                    message="The detected employee sheet contained no usable employee rows.",
                )
            )
            return issues

        duplicate_count = int(dataframe["employee_number"].duplicated().sum())
        if duplicate_count:
            issues.append(
                ValidationIssue(
                    severity="warning",
                    issue_type="duplicate_employee_ids",
                    field_name="employee_number",
                    issue_count=duplicate_count,
                    message="Duplicate employee numbers detected in the snapshot.",
                )
            )

        for field_name in ["work_email", "reporting_manager", "l2_manager", "current_city", "sub_department"]:
            null_count = int(dataframe[field_name].isna().sum()) if field_name in dataframe.columns else len(dataframe)
            if null_count:
                issues.append(
                    ValidationIssue(
                        severity="warning",
                        issue_type="missing_values",
                        field_name=field_name,
                        issue_count=null_count,
                        message=f"{null_count} rows are missing {field_name.replace('_', ' ')}.",
                    )
                )

        malformed_emails = dataframe[
            dataframe["work_email"].notna()
            & ~dataframe["work_email"].astype(str).str.match(EMAIL_PATTERN)
        ]
        if not malformed_emails.empty:
            samples = ", ".join(malformed_emails["work_email"].astype(str).head(5))
            issues.append(
                ValidationIssue(
                    severity="warning",
                    issue_type="malformed_email",
                    field_name="work_email",
                    issue_count=len(malformed_emails),
                    message="Some work email values do not match a standard email pattern.",
                    sample_values=samples,
                )
            )

        for field_name, count in invalid_dates.items():
            if count:
                issues.append(
                    ValidationIssue(
                        severity="warning",
                        issue_type="invalid_date",
                        field_name=field_name,
                        issue_count=count,
                        message=f"{count} rows contained invalid {field_name.replace('_', ' ')} values.",
                    )
                )

        future_exit_mask = (
            dataframe["last_working_day"].notna()
            & dataframe["employment_status"].fillna("").eq("Working")
        )
        future_exit_count = int(future_exit_mask.sum())
        if future_exit_count:
            issues.append(
                ValidationIssue(
                    severity="info",
                    issue_type="pending_exits",
                    field_name="last_working_day",
                    issue_count=future_exit_count,
                    message="Employees marked as Working with a populated last working day will appear in Pending exits.",
                )
            )

        missing_fields = sorted(set(EMPLOYEE_COLUMNS) - available_columns)
        if missing_fields:
            issues.append(
                ValidationIssue(
                    severity="warning",
                    issue_type="schema_gap",
                    field_name="schema",
                    issue_count=len(missing_fields),
                    message="The snapshot is missing canonical columns: " + ", ".join(missing_fields),
                )
            )
        return issues


def determine_compatibility(available_columns: set[str]) -> str:
    if FULL_COMPATIBILITY_FIELDS.issubset(available_columns):
        return "full"
    if COMPATIBLE_REQUIRED_FIELDS.issubset(available_columns):
        missing = set(EMPLOYEE_COLUMNS) - available_columns
        if missing.issubset(OPTIONAL_COMPATIBILITY_FIELDS):
            return "compatible_with_warnings"
    if PARTIAL_MIN_FIELDS.issubset(available_columns):
        return "partial"
    return "rejected"


def parse_as_of_date(
    file_name: str, *, previous_date: date | None = None
) -> tuple[date | None, float, str]:
    match = MONTH_PATTERN.search(file_name)
    if not match:
        return None, 0.0, "Could not parse an as-of date from the filename."

    day = int(match.group("day"))
    month = MONTH_NAME_TO_NUMBER[match.group("month").lower()]
    year_text = match.group("year")
    confidence = 1.0
    note = "As-of date parsed directly from filename."
    if year_text:
        year = int(year_text)
        if year < 100:
            year += 2000
    elif previous_date:
        year = previous_date.year + (1 if month < previous_date.month else 0)
        confidence = 0.75
        note = "As-of year inferred from batch chronology because the filename omitted the year."
    else:
        year = datetime.now().year
        confidence = 0.45
        note = "As-of year defaulted to the current year because the filename omitted the year."

    try:
        return date(year, month, day), confidence, note
    except ValueError:
        return None, 0.0, "Filename contained an invalid calendar date."


def normalize_value(field_name: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if field_name in {"date_joined", "last_working_day", "exit_requested_on"}:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    text = normalize_blankable_text(value)
    if text is None:
        return None
    if field_name in {"employee_number", "work_phone"}:
        return text
    if field_name == "work_email":
        return text.lower()
    if field_name == "gender":
        return text.title()
    if field_name == "employment_status":
        return text.title()
    return normalize_display_value(field_name, text)


def build_status_message(status: str, candidate: SnapshotCandidate) -> str:
    if status == "imported":
        return (
            f"Imported {candidate.row_count} rows from {candidate.source_name} "
            f"as a {candidate.compatibility_level} snapshot."
        )
    if status == "quarantined_duplicate":
        return (
            f"{candidate.source_name} was stored but quarantined because another file for "
            f"{candidate.as_of_date} has a higher row count."
        )
    if status == "already_imported":
        return f"{candidate.source_name} is already present in the archive."
    return f"{candidate.source_name} could not be published: {candidate.parse_note}"


def manual_publish_block_reason(candidate: SnapshotCandidate) -> str | None:
    if candidate.status == "rejected":
        return "Publishing is blocked because the workbook could not be validated against the approved template."
    if candidate.compatibility_level not in MANUAL_UPLOAD_ALLOWED_LEVELS:
        return "Publishing is blocked because mandatory schema fields are missing from the workbook."
    if candidate.as_of_date is None:
        return "Publishing is blocked until a valid snapshot as-of date is confirmed."
    return None


def sanitize_filename(file_name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", file_name)


def resolve_uploaded_by(source_path: Path | None) -> str:
    if source_path is not None and "employee masters" in str(source_path).lower():
        return "fixture_bootstrap"
    try:
        return getuser() or os.environ.get("USERNAME") or "local_user"
    except OSError:
        return os.environ.get("USERNAME") or "local_user"


def _bootstrap_sort_key(path: Path) -> tuple[int, str]:
    prefix_match = re.match(r"(\d+)", path.name)
    prefix = int(prefix_match.group(1)) if prefix_match else 999
    return prefix, path.name.lower()


def _safe_row_values(worksheet: Any, row_index: int) -> tuple[Any, ...] | None:
    try:
        return next(
            worksheet.iter_rows(
                min_row=row_index,
                max_row=row_index,
                values_only=True,
            )
        )
    except StopIteration:
        return None
